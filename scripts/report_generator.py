import os
import pandas as pd
from fpdf import FPDF
import openpyxl

def clean_risk_text(risk_level):
    """Remove emojis from risk level descriptions."""
    return risk_level.replace("🔴", "High Risk:").replace("🟠", "Medium Risk:").replace("🟢", "Low Risk:")

def generate_pdf_report(project_name, labor_cost, material_cost, equipment_cost, miscellaneous_cost, duration, predicted_cost, risk_level):
    """Generate a PDF report with project details and predictions."""
    
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    
    # Title
    pdf.cell(200, 10, "Project Cost Estimation Report", ln=True, align="C")
    pdf.ln(10)

    # Project Details
    pdf.set_font("Arial", "", 12)
    pdf.cell(200, 10, f"Project Name: {project_name}", ln=True)
    pdf.cell(200, 10, f"Duration: {duration} months", ln=True)
    pdf.ln(5)

    # Cost Breakdown
    pdf.cell(200, 10, "Cost Breakdown:", ln=True)
    pdf.cell(200, 10, f"Labor Cost: INR {labor_cost}", ln=True)
    pdf.cell(200, 10, f"Material Cost: INR {material_cost}", ln=True)
    pdf.cell(200, 10, f"Equipment Cost: INR {equipment_cost}", ln=True)
    pdf.cell(200, 10, f"Miscellaneous Costs: INR {miscellaneous_cost}", ln=True)
    pdf.ln(5)

    # Prediction & Risk
    pdf.cell(200, 10, f"Predicted Total Cost: INR {predicted_cost:.2f}", ln=True)
    pdf.cell(200, 10, f"Risk Level: {clean_risk_text(risk_level)}", ln=True)

    # Ensure "reports" directory exists
    os.makedirs("reports", exist_ok=True)

    # Save the PDF
    pdf_file = f"reports/{project_name}_cost_report.pdf"
    pdf.output(pdf_file, "F")
    return pdf_file

def generate_excel_report(project_name, labor_cost, material_cost, equipment_cost, miscellaneous_cost, duration, predicted_cost, risk_level):
    """Generate an Excel report with project details and predictions."""
    
    # Create a DataFrame
    data = {
        "Project Name": [project_name],
        "Duration (months)": [duration],
        "Labor Cost (INR)": [labor_cost],
        "Material Cost (INR)": [material_cost],
        "Equipment Cost (INR)": [equipment_cost],
        "Miscellaneous Cost (INR)": [miscellaneous_cost],
        "Predicted Total Cost (INR)": [predicted_cost],
        "Risk Level": [clean_risk_text(risk_level)]
    }

    df = pd.DataFrame(data)

    # Ensure "reports" directory exists
    os.makedirs("reports", exist_ok=True)

    # Save to Excel
    excel_file = f"reports/{project_name}_cost_report.xlsx"
    df.to_excel(excel_file, index=False)
    return excel_file
